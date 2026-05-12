from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from sqlalchemy import func
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from urllib.parse import quote
from app.database import get_db
from app.models import Submission, Event

router = APIRouter()
templates = Jinja2Templates(directory="templates")


def _valid_filter(q, event_id):
    """정상 데이터 기본 필터 (NULL 포함 처리)"""
    return q.filter(
        Submission.event_id == event_id,
        Submission.is_excluded.isnot(True),
        Submission.is_error.isnot(True),
    )


def get_stats(event_id: int, db: Session) -> dict:
    total_count = _valid_filter(db.query(Submission), event_id).count()

    item1_stats = (
        _valid_filter(db.query(Submission.item1, func.count(Submission.id).label("cnt")), event_id)
        .group_by(Submission.item1)
        .order_by(func.count(Submission.id).desc())
        .all()
    )
    item2_stats = (
        _valid_filter(db.query(Submission.item2, func.count(Submission.id).label("cnt")), event_id)
        .group_by(Submission.item2)
        .order_by(func.count(Submission.id).desc())
        .all()
    )
    item3_stats = (
        _valid_filter(db.query(Submission.item3, func.count(Submission.id).label("cnt")), event_id)
        .group_by(Submission.item3)
        .order_by(func.count(Submission.id).desc())
        .all()
    )
    combo_stats = (
        _valid_filter(
            db.query(Submission.item1, Submission.item2, func.count(Submission.id).label("cnt")), event_id
        )
        .group_by(Submission.item1, Submission.item2)
        .order_by(func.count(Submission.id).desc())
        .all()
    )

    error_count = db.query(Submission).filter(
        Submission.event_id == event_id,
        Submission.is_error == True,
    ).count()
    unresolved_count = db.query(Submission).filter(
        Submission.event_id == event_id,
        Submission.is_error == True,
        Submission.is_resolved.isnot(True),
    ).count()

    return {
        "total_count": total_count,
        "item1_stats": item1_stats,
        "item2_stats": item2_stats,
        "item3_stats": item3_stats,
        "combo_stats": combo_stats,
        "error_count": error_count,
        "unresolved_count": unresolved_count,
    }


@router.get("/events/{event_id}/statistics", response_class=HTMLResponse)
async def statistics(
    request: Request,
    event_id: int,
    q_name: str = "",
    q_emp: str = "",
    q_item: str = "",
    tab: str = "chart",   # chart | grid
    page: int = 1,
    db: Session = Depends(get_db),
):
    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise HTTPException(status_code=404)

    stats = get_stats(event_id, db)

    # 그리드 쿼리 (정상 데이터 + 검색 조건)
    PAGE_SIZE = 50
    grid_q = _valid_filter(db.query(Submission), event_id)

    if q_name.strip():
        grid_q = grid_q.filter(Submission.name.contains(q_name.strip()))
    if q_emp.strip():
        grid_q = grid_q.filter(Submission.employee_id.contains(q_emp.strip()))
    if q_item.strip():
        kw = q_item.strip()
        grid_q = grid_q.filter(
            Submission.item1.contains(kw) |
            Submission.item2.contains(kw) |
            Submission.item3.contains(kw)
        )

    grid_total = grid_q.count()
    grid_data = (
        grid_q.order_by(Submission.employee_id)
        .offset((page - 1) * PAGE_SIZE)
        .limit(PAGE_SIZE)
        .all()
    )
    grid_total_pages = max(1, (grid_total + PAGE_SIZE - 1) // PAGE_SIZE)

    return templates.TemplateResponse("statistics.html", {
        "request": request,
        "event": event,
        "active_tab": tab,
        "q_name": q_name,
        "q_emp": q_emp,
        "q_item": q_item,
        "page": page,
        "grid_data": grid_data,
        "grid_total": grid_total,
        "grid_total_pages": grid_total_pages,
        **stats,
    })


@router.get("/events/{event_id}/export")
async def export_excel(event_id: int, db: Session = Depends(get_db)):
    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise HTTPException(status_code=404)

    stats = get_stats(event_id, db)
    wb = openpyxl.Workbook()

    h_font = Font(bold=True, color="FFFFFF")
    h_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def style_header(ws, headers):
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = h_font; c.fill = h_fill; c.alignment = center; c.border = thin

    ws1 = wb.active
    ws1.title = "전체 데이터"
    style_header(ws1, ["사번", "이름", "전화번호", "취합사항1", "취합사항2", "취합사항3"])
    subs = _valid_filter(db.query(Submission), event_id).order_by(Submission.employee_id).all()
    for r, s in enumerate(subs, 2):
        for c, v in enumerate([s.employee_id, s.name, s.phone, s.item1, s.item2, s.item3], 1):
            cell = ws1.cell(row=r, column=c, value=v)
            cell.border = thin

    ws2 = wb.create_sheet("취합사항1 통계")
    style_header(ws2, ["취합사항1", "인원수"])
    for r, (item, cnt) in enumerate(stats["item1_stats"], 2):
        ws2.cell(r, 1, item or "미입력"); ws2.cell(r, 2, cnt)

    ws3 = wb.create_sheet("취합사항2 통계")
    style_header(ws3, ["취합사항2", "인원수"])
    for r, (item, cnt) in enumerate(stats["item2_stats"], 2):
        ws3.cell(r, 1, item or "미입력"); ws3.cell(r, 2, cnt)

    ws4 = wb.create_sheet("항목 조합 통계")
    style_header(ws4, ["취합사항1", "취합사항2", "인원수"])
    for r, (i1, i2, cnt) in enumerate(stats["combo_stats"], 2):
        ws4.cell(r, 1, i1 or "-"); ws4.cell(r, 2, i2 or "-"); ws4.cell(r, 3, cnt)

    for ws in [ws1, ws2, ws3, ws4]:
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    safe_name = quote(f"{event.name}_통계.xlsx")
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{safe_name}"},
    )
