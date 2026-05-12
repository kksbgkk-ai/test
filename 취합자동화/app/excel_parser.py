import openpyxl
import os
import shutil
from datetime import datetime
from sqlalchemy.orm import Session
from app.models import Submission, SourceFile
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config import EXCEL_COLUMN_MAP, PROCESSED_FOLDER, EXCEL_DATA_START_ROW


def _build_col_index(ws) -> dict:
    """1행 헤더를 읽어 {field: 열번호} 매핑 반환"""
    header_map = {}
    for cell in ws[1]:
        if cell.value is not None:
            header_map[str(cell.value).strip()] = cell.column

    col_index = {}
    for field, header in EXCEL_COLUMN_MAP.items():
        col_index[field] = header_map.get(header)
    return col_index


def _cell_val(ws, row: int, col):
    """셀 값 안전하게 읽기 (None/공백 처리)"""
    if col is None:
        return None
    val = ws.cell(row=row, column=col).value
    if val is None:
        return None
    if isinstance(val, str):
        val = val.strip()
        return val if val else None
    return str(val).strip()


def _get_existing_employee_ids(event_id: int, db: Session) -> set:
    """DB에 이미 저장된 사번 목록 (정상 데이터만)"""
    rows = db.query(Submission.employee_id).filter(
        Submission.event_id == event_id,
        Submission.is_excluded.isnot(True),
    ).all()
    return {r[0] for r in rows}


def parse_excel_file(filepath: str, event_id: int, db: Session) -> dict:
    """
    엑셀 파일 파싱 → DB 저장.
    반환: {success, row_count, error_count, message}

    중복 감지 전략:
    - DB에 이미 있는 사번: 파일 처리 시작 전 일괄 로드 (다른 파일 중복)
    - 같은 파일 내 중복: Python set으로 추적 (autoflush=False 문제 우회)
    """
    filename = os.path.basename(filepath)

    source_file = SourceFile(
        event_id=event_id,
        filename=filename,
        filepath=filepath,
        status="processing",
    )
    db.add(source_file)
    db.commit()
    db.refresh(source_file)

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active

        col_index = _build_col_index(ws)

        if col_index.get("employee_id") is None:
            raise ValueError(f"'사번' 컬럼을 찾을 수 없습니다. (파일: {filename})")

        # 파일 처리 전: DB에 이미 존재하는 사번 목록 로드
        db_existing_ids = _get_existing_employee_ids(event_id, db)
        # 이번 파일에서 이미 처리한 사번 (파일 내 중복 추적용)
        file_seen_ids: set[str] = set()

        row_count = 0
        error_count = 0

        for row in range(EXCEL_DATA_START_ROW, ws.max_row + 1):
            employee_id = _cell_val(ws, row, col_index.get("employee_id"))

            if not employee_id:
                continue

            row_count += 1

            # 중복 여부 판단
            in_db = employee_id in db_existing_ids        # 다른 파일에서 이미 들어온 사번
            in_file = employee_id in file_seen_ids         # 이번 파일 내 앞에서 이미 나온 사번
            is_duplicate = in_db or in_file
            is_error = is_duplicate
            error_reason = None

            if is_duplicate:
                error_count += 1
                if in_file:
                    error_reason = f"파일 내 사번 중복: {employee_id}"
                else:
                    error_reason = f"사번 중복 (다른 파일에 존재): {employee_id}"

                # DB에 이미 있는 기존 데이터도 중복 표시
                if in_db:
                    existing = db.query(Submission).filter(
                        Submission.event_id == event_id,
                        Submission.employee_id == employee_id,
                        Submission.is_excluded.isnot(True),
                        Submission.is_error.isnot(True),
                    ).first()
                    if existing:
                        existing.is_duplicate = True
                        existing.is_error = True
                        existing.error_reason = f"사번 중복 (파일: {filename})"

            sub = Submission(
                event_id=event_id,
                source_file_id=source_file.id,
                employee_id=employee_id,
                name=_cell_val(ws, row, col_index.get("name")) or "",
                phone=_cell_val(ws, row, col_index.get("phone")),
                item1=_cell_val(ws, row, col_index.get("item1")),
                item2=_cell_val(ws, row, col_index.get("item2")),
                item3=_cell_val(ws, row, col_index.get("item3")),
                is_duplicate=is_duplicate,
                is_error=is_error,
                error_reason=error_reason,
                is_excluded=is_duplicate,  # 중복은 기본 통계 제외
            )
            db.add(sub)

            # 정상 데이터만 seen에 추가 (중복은 추가 안 함 — 원본 보존)
            if not is_duplicate:
                file_seen_ids.add(employee_id)

        db.commit()

        # 처리 완료 폴더로 이동
        os.makedirs(PROCESSED_FOLDER, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        dest = os.path.join(PROCESSED_FOLDER, f"{timestamp}_{filename}")
        shutil.move(filepath, dest)

        source_file.row_count = row_count
        source_file.error_count = error_count
        source_file.status = "ok"
        db.commit()

        return {
            "success": True,
            "row_count": row_count,
            "error_count": error_count,
            "message": f"처리 완료: {row_count}건 (중복/오류: {error_count}건)",
        }

    except Exception as e:
        db.rollback()
        source_file.status = "error"
        source_file.error_message = str(e)
        db.commit()
        return {
            "success": False,
            "row_count": 0,
            "error_count": 0,
            "message": f"오류: {str(e)}",
        }


def get_existing_filenames(event_id: int, db: Session) -> set:
    """이미 처리된 파일명 목록"""
    files = db.query(SourceFile).filter(SourceFile.event_id == event_id).all()
    return {os.path.basename(f.filename) for f in files}
